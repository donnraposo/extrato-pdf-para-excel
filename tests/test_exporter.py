from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from extrato_parser.exporter import export_to_excel
from extrato_parser.models import ParseResult, StatementEntry


def test_excel_has_expected_sheets_and_numeric_values(tmp_path):
    result = ParseResult(Path("origem.pdf"))
    result.entries.append(StatementEntry(date(2026, 2, 5), "PIX", None, Decimal("100.50"), None, 1, "PIX"))
    output = tmp_path / "saida.xlsx"
    export_to_excel(result, output)
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["Lançamentos", "Conferência", "Metadados"]
    assert workbook["Lançamentos"]["F2"].value == 100.5
    assert workbook["Lançamentos"]["D1"].value == "Crédito (R$)"
    assert workbook["Lançamentos"]["E1"].value == "Débito (R$)"


def test_excel_uses_dynamic_itau_columns(tmp_path):
    result = ParseResult(Path("itau.pdf"))
    result.output_fields = ["date", "description", "credit", "debit", "movement", "balance"]
    result.field_labels = {"credit": "Entradas (R$)", "debit": "Saídas (R$)"}
    result.entries.append(StatementEntry(date(2026, 1, 2), "Sispag", None, Decimal("2990"), None, 1, "Sispag", credit=Decimal("2990")))
    output = tmp_path / "itau.xlsx"
    export_to_excel(result, output)
    sheet = load_workbook(output)["Lançamentos"]
    assert [cell.value for cell in sheet[1]][:6] == ["Data", "Descrição", "Entradas (R$)", "Saídas (R$)", "Movimento (R$)", "Saldo (R$)"]
    assert "Nº Documento" not in [cell.value for cell in sheet[1]]


def test_excel_exports_caixa_effective_date_as_date_without_time(tmp_path):
    result = ParseResult(Path("caixa.pdf"))
    result.output_fields = ["date", "effective_date", "document", "description", "movement", "balance"]
    result.field_labels = {"document": "Documento", "description": "Histórico", "movement": "Valor (R$)"}
    result.entries.append(
        StatementEntry(
            date(2026, 1, 5),
            "PIX RECEBIDO",
            "051452",
            Decimal("2625.00"),
            "15.187,34 D",
            1,
            "linha original",
            effective_date=datetime(2026, 1, 5, 14, 52),
            credit=Decimal("2625.00"),
        )
    )
    output = tmp_path / "caixa.xlsx"
    export_to_excel(result, output)
    sheet = load_workbook(output)["Lançamentos"]
    assert [cell.value for cell in sheet[1]][:4] == ["Data", "Data Efetiva", "Documento", "Histórico"]
    assert [cell.value for cell in sheet[1]][4:6] == ["Valor (R$)", "Saldo (R$)"]
    assert sheet["B2"].value == datetime(2026, 1, 5, 14, 52)
    assert sheet["B2"].number_format == "dd/mm/yyyy hh:mm"
    assert sheet["F2"].value == "15.187,34 D"
