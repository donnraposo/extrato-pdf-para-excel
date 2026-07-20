from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ParseResult


DEFAULT_FIELDS = ["date", "description", "document", "credit", "debit", "movement", "balance"]
DEFAULT_LABELS = {
    "date": "Data",
    "effective_date": "Data Efetiva",
    "description": "Descrição",
    "document": "Nº Documento",
    "credit": "Crédito (R$)",
    "debit": "Débito (R$)",
    "movement": "Movimento (R$)",
    "balance": "Saldo (R$)",
}
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def _style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for column in sheet.columns:
        values = [len(str(cell.value or "")) for cell in column]
        width = min(max(values + [10]) + 2, 60)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def export_to_excel(result: ParseResult, output: Path) -> None:
    workbook = Workbook()
    entries_sheet = workbook.active
    entries_sheet.title = "Lançamentos"
    extra_headers = list(result.detected_extra_fields)
    output_fields = result.output_fields or DEFAULT_FIELDS
    labels = {**DEFAULT_LABELS, **result.field_labels}
    headers = [labels[field] for field in output_fields] + extra_headers + ["Página de origem", "Status"]
    entries_sheet.append(headers)

    for entry in result.entries:
        values = {
            "date": entry.transaction_date,
            "effective_date": entry.effective_date,
            "description": entry.description,
            "document": entry.document_number,
            "credit": float(entry.credit) if entry.credit is not None else None,
            "debit": float(entry.debit) if entry.debit is not None else None,
            "movement": float(entry.movement) if entry.movement is not None else None,
            "balance": float(entry.balance) if isinstance(entry.balance, Decimal) else entry.balance,
        }
        row = [values[field] for field in output_fields]
        row.extend(entry.extra_fields.get(name) for name in extra_headers)
        row.extend([entry.page, "Conferir" if entry.needs_review else "OK"])
        entries_sheet.append(row)
        if "date" in output_fields:
            entries_sheet.cell(entries_sheet.max_row, output_fields.index("date") + 1).number_format = "dd/mm/yyyy"
        if "effective_date" in output_fields:
            entries_sheet.cell(entries_sheet.max_row, output_fields.index("effective_date") + 1).number_format = "dd/mm/yyyy hh:mm"
        for field in ("credit", "debit", "movement", "balance"):
            if field in output_fields:
                entries_sheet.cell(entries_sheet.max_row, output_fields.index(field) + 1).number_format = '#,##0.00;[Red]-#,##0.00'
    _style_sheet(entries_sheet)

    review = workbook.create_sheet("Conferência")
    review.append(["Página", "Tipo", "Descrição/Texto original", "Motivo"])
    for entry in result.entries:
        if entry.needs_review:
            review.append([entry.page, "Lançamento", entry.source_text, "; ".join(entry.warnings) or "Dados incompletos"])
    for page, text, reason in result.rejected_lines:
        review.append([page, "Linha não estruturada", text, reason])
    _style_sheet(review)

    metadata = workbook.create_sheet("Metadados")
    metadata.append(["Campo", "Valor"])
    metadata.append(["Arquivo de origem", result.source.name])
    metadata.append(["Processado em", datetime.now().astimezone().isoformat(timespec="seconds")])
    for key, value in result.metadata.items():
        metadata.append([key, value])
    _style_sheet(metadata)

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f".{output.name}.tmp.xlsx")
    try:
        workbook.save(temporary)
        temporary.replace(output)
    finally:
        if temporary.exists():
            temporary.unlink()
